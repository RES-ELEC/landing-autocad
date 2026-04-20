"""
Scraper entreprises - recherche-entreprises.api.gouv.fr
SARL / SAS / EURL / SASU actives pour un code APE + département donné.

Usage : python scrape_inpi.py
"""

import requests, time, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

API_URL = "https://recherche-entreprises.api.gouv.fr/search"
DEPARTEMENT = "91"
CODE_APE = "71.12B"
PER_PAGE = 25
CHUNK_SIZE = 1500

# Anti rate-limit
DELAY_BETWEEN_PAGES = 1.0   # secondes entre chaque page
PAUSE_EVERY = 20            # toutes les N pages, longue pause
PAUSE_DURATION = 15         # durée de la longue pause (s)
MAX_RETRIES = 5

HEADERS = {
    "User-Agent": "RES-ELEC-Prospection/1.0 (contact@reselec.fr)",
    "Accept": "application/json",
}

OUTPUT_FILE = rf"C:\Users\chapo\Desktop\Landing RES ELEC\entreprises_{CODE_APE.replace('.', '')}_{DEPARTEMENT}.xlsx"

CODES_SARL_SAS = {"5498", "5499", "5710", "5720"}
LIBELLES_NJ = {"5498": "EURL", "5499": "SARL", "5710": "SAS", "5720": "SASU"}


def fetch_page(page, max_retries=MAX_RETRIES):
    params = {
        "q": "",
        "activite_principale": CODE_APE,
        "departement": DEPARTEMENT,
        "etat_administratif": "A",
        "page": page,
        "per_page": PER_PAGE,
    }
    for attempt in range(max_retries):
        try:
            resp = requests.get(API_URL, params=params, headers=HEADERS, timeout=30)
        except requests.exceptions.RequestException as e:
            wait = 10 * (attempt + 1)
            print(f"\n  ⚠️  Erreur réseau ({e}), attente {wait}s...", flush=True)
            time.sleep(wait)
            continue

        if resp.status_code == 200:
            return resp.json()

        if resp.status_code == 429:
            retry_after = 0
            try:
                retry_after = int(resp.headers.get("Retry-After", 0))
            except (TypeError, ValueError):
                retry_after = 0
            # backoff exponentiel : 30, 60, 120, 240, 480
            wait = max(retry_after, 30 * (2 ** attempt))
            print(f"\n  ⏳ 429 Rate limit — attente {wait}s (tentative {attempt+1}/{max_retries})...", flush=True)
            time.sleep(wait)
            continue

        # autre erreur HTTP : on arrête direct
        resp.raise_for_status()

    raise requests.exceptions.RequestException(
        f"Échec page {page} après {max_retries} tentatives (429 persistant)"
    )


def extract_dirigeant(company):
    d = (company.get("dirigeants") or [None])[0]
    if not d:
        return ""
    prenom = d.get("prenoms", "") or d.get("prenom", "") or ""
    nom = d.get("nom", "") or ""
    if prenom or nom:
        return f"{prenom} {nom}".strip().title()
    return d.get("denomination", "") or ""


def scrape_all():
    print("Récupération page 1...")
    data = fetch_page(1)
    total, total_pages = data.get("total_results", 0), data.get("total_pages", 1)
    print(f"  → {total} entreprises trouvées ({total_pages} pages)\n")
    if total == 0:
        return []

    companies = []
    for page in range(1, total_pages + 1):
        if page > 1:
            time.sleep(DELAY_BETWEEN_PAGES)
            if (page - 1) % PAUSE_EVERY == 0:
                print(f"\n  💤 Pause {PAUSE_DURATION}s (anti rate-limit)...", flush=True)
                time.sleep(PAUSE_DURATION)
            data = fetch_page(page)

        for c in data.get("results", []):
            nj = str(c.get("nature_juridique", "") or "")
            if nj not in CODES_SARL_SAS:
                continue
            s = c.get("siege", {}) or {}
            companies.append({
                "denomination": c.get("nom_complet", "") or "",
                "siren": c.get("siren", "") or "",
                "dirigeant": extract_dirigeant(c),
                "forme_juridique": LIBELLES_NJ.get(nj, nj),
                "code_postal": s.get("code_postal", "") or "",
                "ville": s.get("libelle_commune", "") or "",
                "departement": s.get("departement", "") or DEPARTEMENT,
            })
        done = min(page * PER_PAGE, total)
        print(f"  Page {page}/{total_pages} — {done}/{total} — {len(companies)} SARL/SAS", flush=True)

    print(f"\n✅ {len(companies)} entreprises SARL/SAS retenues sur {total}")
    return companies


def export_excel(companies, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Entreprises {CODE_APE} - {DEPARTEMENT}"
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1B3A5C")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    bd = Border(left=Side("thin", "D0D0D0"), right=Side("thin", "D0D0D0"),
                top=Side("thin", "D0D0D0"), bottom=Side("thin", "D0D0D0"))
    alt = PatternFill("solid", fgColor="F2F6FA")
    headers = ["N°", "Dénomination", "Dirigeant", "SIREN", "Forme juridique", "Code postal", "Ville", "Département"]
    widths = [6, 40, 30, 15, 18, 14, 25, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, bd
        ws.column_dimensions[c.column_letter].width = w
    for ri, comp in enumerate(companies, 2):
        vals = [ri - 1, comp["denomination"], comp["dirigeant"], comp["siren"],
                comp["forme_juridique"], comp["code_postal"], comp["ville"], comp["departement"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.border = cf, bd
            if ri % 2 == 0:
                c.fill = alt
            if ci == 4:
                c.number_format = "@"
    ws.auto_filter.ref = f"A1:H{len(companies) + 1}"
    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"📄 Fichier exporté : {filename}")


def main():
    print("=" * 60)
    print(f"  SCRAPER ENTREPRISES - APE {CODE_APE} / Dept {DEPARTEMENT}")
    print("  SARL + SAS + EURL + SASU actives")
    print("=" * 60, "\n")
    try:
        companies = scrape_all()
    except requests.exceptions.RequestException as e:
        print(f"\n❌ Erreur : {e}")
        sys.exit(1)
    if not companies:
        print("Aucune entreprise trouvée.")
        sys.exit(0)

    total = len(companies)
    nb_chunks = (total + CHUNK_SIZE - 1) // CHUNK_SIZE
    base, ext = OUTPUT_FILE.rsplit(".", 1)
    for i in range(nb_chunks):
        chunk = companies[i * CHUNK_SIZE:(i + 1) * CHUNK_SIZE]
        fname = OUTPUT_FILE if nb_chunks == 1 else f"{base}_part{i + 1}.{ext}"
        export_excel(chunk, fname)
    print(f"\nTerminé ! {total} lignes au total ({nb_chunks} fichier(s))")


if __name__ == "__main__":
    main()
