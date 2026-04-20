"""
Enrichissement d'un ou plusieurs fichiers Excel d'entreprises avec leur site web.

Mode MAX FIABILITÉ :
  - 2 moteurs (DuckDuckGo HTML + Qwant)
  - 4 formulations de requête par société
  - Jusqu'à 6 candidats validés par société
  - Validation multi-pages : homepage + /mentions-legales + /contact
  - Scoring strict avec SIREN / nom / dirigeant / slug domaine
  - Seuil de confiance élevé

Produit :
  - Un fichier <nom>_avecsite.xlsx pour chaque source (sauvegarde incrémentale)
  - Un fichier unique FUSIONNÉ : entreprises_MERGED_avecsite.xlsx
    dans le dossier du PREMIER fichier source fourni

Usage :
    python find_websites.py

    → une boîte de dialogue s'ouvre pour sélectionner le(s) fichier(s) Excel
    → Ctrl+clic ou Maj+clic pour sélectionner plusieurs fichiers d'un coup
"""

import glob
import os
import random
import re
import sys
import time
import tkinter as tk
import unicodedata
from tkinter import filedialog, messagebox
from urllib.parse import unquote, urlparse

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- Config MAX FIABILITÉ -------------------------------------------------

DELAY_MIN = 3.0            # délai mini entre sociétés (s)
DELAY_MAX = 5.5            # délai maxi entre sociétés (s)
PAUSE_EVERY = 25           # longue pause toutes les N sociétés
PAUSE_DURATION = 30        # durée de la longue pause (s)
SAVE_EVERY = 20            # sauvegarde intermédiaire
SEARCH_TIMEOUT = 15
FETCH_TIMEOUT = 10
MAX_CANDIDATES_TO_VALIDATE = 6    # candidats évalués par société
SCORE_THRESHOLD = 45              # seuil minimum pour accepter
DEFINITIVE_SCORE = 150            # score qui déclenche l'arrêt (SIREN trouvé)

# Domaines à ignorer
BLACKLIST_DOMAINS = {
    "societe.com", "verif.com", "pappers.fr", "infogreffe.fr", "bodacc.fr",
    "score3.fr", "bilansgratuits.fr", "manageo.fr", "kompass.com", "kompass.fr",
    "pagesjaunes.fr", "pages-jaunes.fr", "118712.fr", "118000.fr", "118218.fr",
    "linkedin.com", "facebook.com", "instagram.com", "twitter.com", "x.com",
    "youtube.com", "tiktok.com", "pinterest.fr", "pinterest.com",
    "indeed.fr", "indeed.com", "glassdoor.fr", "welcometothejungle.com", "hellowork.com",
    "leboncoin.fr", "mappy.com", "google.com", "google.fr", "bing.com",
    "wikipedia.org", "wikiwand.com", "fr.wikipedia.org",
    "data.gouv.fr", "insee.fr", "sirene.fr", "annuaire-entreprises.data.gouv.fr",
    "europages.fr", "europages.com", "trouverundevis.com", "hellopro.fr",
    "batiactu.com", "lemoniteur.fr", "lefigaro.fr", "lesechos.fr",
    "yelp.fr", "yelp.com", "tripadvisor.fr", "justacote.com",
    "duckduckgo.com", "qwant.com", "ecosia.org", "startpage.com",
    "dirigeant.fr", "creationdentreprise.fr", "usine-digitale.fr",
    "b-reputation.com", "b-reputation.fr", "scoresandmore.fr",
    "go.mail.ru", "youscribe.com", "scribd.com", "doc-etudiant.fr",
    "journaldunet.com", "journaldunet.fr",
    "corporama.com", "dnb.com", "opencorporates.com", "northdata.com",
    "amazon.fr", "amazon.com", "cdiscount.com", "ebay.fr",
    "allo-annuaire.com", "allo-pro.com", "annuaire-gratuit-pro.com",
    "trustpilot.fr", "trustpilot.com", "avis-verifies.com",
    "verif-siren.com", "score-entreprise.com", "societeinfo.com",
    "entreprises.lefigaro.fr", "bfmbusiness.bfmtv.com",
}

STOPWORDS = {
    "sarl", "sas", "sasu", "eurl", "sa", "sci", "scop", "selarl", "snc",
    "le", "la", "les", "de", "du", "des", "et", "en", "au", "aux",
    "l", "d", "un", "une", "the", "and", "or", "for",
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

DDG_HTML_URL = "https://html.duckduckgo.com/html/"
QWANT_URL = "https://www.qwant.com/"

# Sous-pages à tester quand la home ne suffit pas
VALIDATION_SUBPATHS = [
    "mentions-legales", "mentions-legales/", "mentionslegales",
    "contact", "contact/", "nous-contacter",
    "qui-sommes-nous", "a-propos", "about",
]


# --- Normalisation -------------------------------------------------------

def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normalize(s: str) -> str:
    s = strip_accents(s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())


def slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize(s))


def significant_words(name: str) -> list:
    words = normalize(name).split()
    return [w for w in words if len(w) >= 3 and w not in STOPWORDS]


def clean_query_name(name: str) -> str:
    if not name:
        return ""
    # Retire les contenus entre parenthèses (souvent noms dupliqués INPI)
    name = re.sub(r"\([^)]*\)", " ", name)
    # Retire ponctuation type S.A.R.L. avant le match des formes
    name = re.sub(r"\bS\.?A\.?R\.?L\.?\b", " ", name, flags=re.I)
    name = re.sub(r"\b(SARL|SAS|SASU|EURL|SA|SCI|SCOP|SELARL|SNC)\b", " ", name, flags=re.I)
    # Retire ponctuation résiduelle
    name = re.sub(r"[.,;:/\\-]+", " ", name)
    return " ".join(name.split()).strip()


def _domain_of(url: str) -> str:
    try:
        netloc = urlparse(url).netloc.lower()
        if netloc.startswith("www."):
            netloc = netloc[4:]
        return netloc
    except Exception:
        return ""


def is_blacklisted(url: str) -> bool:
    d = _domain_of(url)
    if not d:
        return True
    for bad in BLACKLIST_DOMAINS:
        if d == bad or d.endswith("." + bad):
            return True
    return False


# --- Recherche web --------------------------------------------------------

def _extract_ddg_links(html: str):
    links = []
    for m in re.finditer(r'href="(?:/l/\?[^"]*?uddg=)([^"&]+)', html):
        try:
            decoded = unquote(m.group(1))
            if decoded.startswith("http"):
                links.append(decoded)
        except Exception:
            pass
    for m in re.finditer(r'class="result__a"[^>]*href="([^"]+)"', html):
        url = m.group(1)
        if url.startswith("http"):
            links.append(url)
    return links


def _extract_qwant_links(html: str):
    links = []
    for m in re.finditer(r'<a[^>]+href="(https?://[^"]+)"', html):
        url = m.group(1)
        if "qwant.com" not in url and "qwantjunior" not in url:
            links.append(url)
    return links


def search_ddg(query: str, session: requests.Session):
    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "fr-FR,fr;q=0.9",
    }
    try:
        resp = session.post(DDG_HTML_URL, data={"q": query, "kl": "fr-fr"},
                            headers=headers, timeout=SEARCH_TIMEOUT)
    except requests.exceptions.RequestException:
        return None
    if resp.status_code == 200:
        return _extract_ddg_links(resp.text)
    if resp.status_code in (202, 429):
        return "RATELIMIT"
    return None


def search_qwant(query: str, session: requests.Session):
    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "fr-FR,fr;q=0.9",
    }
    try:
        resp = session.get(QWANT_URL, params={"q": query, "l": "fr"},
                           headers=headers, timeout=SEARCH_TIMEOUT)
    except requests.exceptions.RequestException:
        return None
    if resp.status_code == 200:
        return _extract_qwant_links(resp.text)
    if resp.status_code in (202, 429):
        return "RATELIMIT"
    return None


def collect_candidates(denomination, ville, cp, session):
    name = clean_query_name(denomination)
    if not name:
        return []

    queries = []
    if ville:
        queries.append(f'"{name}" {ville}')
        queries.append(f'{name} {ville} contact')
    if cp:
        queries.append(f'{name} {cp}')
    queries.append(f'"{name}"')

    all_links = []
    seen_domains = set()

    for q in queries:
        links = search_ddg(q, session)
        if links == "RATELIMIT":
            time.sleep(25)
            links = search_ddg(q, session)
        if links:
            for url in links:
                if is_blacklisted(url):
                    continue
                d = _domain_of(url)
                if d and d not in seen_domains:
                    seen_domains.add(d)
                    all_links.append(url)
            if len(all_links) >= MAX_CANDIDATES_TO_VALIDATE:
                break

        time.sleep(random.uniform(1.2, 2.2))

    # Fallback Qwant si peu de résultats
    if len(all_links) < 2:
        q = f'"{name}" {ville}' if ville else f'"{name}"'
        links = search_qwant(q, session)
        if links and links != "RATELIMIT":
            for url in links:
                if is_blacklisted(url):
                    continue
                d = _domain_of(url)
                if d and d not in seen_domains:
                    seen_domains.add(d)
                    all_links.append(url)

    return all_links[:MAX_CANDIDATES_TO_VALIDATE]


# --- Validation par fetch -------------------------------------------------

def fetch_url(url: str, session: requests.Session) -> str:
    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "fr-FR,fr;q=0.9",
    }
    try:
        resp = session.get(url, headers=headers, timeout=FETCH_TIMEOUT,
                           allow_redirects=True)
        if resp.status_code == 200 and resp.text:
            return resp.text
    except requests.exceptions.RequestException:
        pass
    return ""


def strip_html(html: str) -> str:
    html = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.DOTALL | re.I)
    html = re.sub(r"<style[^>]*>.*?</style>", " ", html, flags=re.DOTALL | re.I)
    return re.sub(r"<[^>]+>", " ", html)


def score_from_text(plain_text: str, denomination: str, siren: str,
                    dirigeant: str, domain: str) -> int:
    """
    Scoring strict — un vrai site d'entreprise doit avoir SOIT le SIREN,
    SOIT une corrélation avec le nom de domaine. Le texte seul ne suffit pas.
    """
    if not plain_text:
        return 0

    text_norm = normalize(plain_text)
    score = 0

    # =========================================================
    # 1) SIREN présent → match certain (termine la recherche)
    # =========================================================
    if siren:
        siren_clean = re.sub(r"\s+", "", str(siren))
        if siren_clean and len(siren_clean) == 9:
            raw_text = re.sub(r"\s+", "", plain_text)
            if siren_clean in raw_text:
                score += 200

    # =========================================================
    # 2) Corrélation NOM <-> DOMAINE (signal principal)
    #    Un vrai site a presque toujours le nom dans son domaine.
    # =========================================================
    domain_root = domain.split(".")[0] if domain else ""
    domain_slug = re.sub(r"[^a-z0-9]+", "", domain_root)
    name_clean = clean_query_name(denomination)
    name_slug = slug(name_clean)

    domain_bonus = 0
    if name_slug and domain_slug:
        # a) slug complet du nom dans le domaine (ou l'inverse)
        if name_slug in domain_slug or domain_slug in name_slug:
            domain_bonus = 55
        else:
            # b) au moins un mot significatif distinctif (>= 4 chars) dans le domaine
            distinctive_words = [w for w in significant_words(denomination) if len(w) >= 4]
            matched = [w for w in distinctive_words if w in domain_slug]
            if matched:
                # plus il y en a, plus c'est solide
                if len(matched) >= 2:
                    domain_bonus = 55
                else:
                    domain_bonus = 45
            else:
                # c) préfixe de 5+ caractères
                if len(name_slug) >= 5 and name_slug[:5] in domain_slug:
                    domain_bonus = 15
    score += domain_bonus

    # =========================================================
    # 3) Signaux textuels — faibles, utiles seulement en appoint
    # =========================================================
    # Nom complet dans le texte
    if name_clean:
        name_norm = normalize(name_clean)
        if len(name_norm) >= 5 and name_norm in text_norm:
            score += 15

    # Mots distinctifs (>= 4 chars) dans le texte
    distinctive = [w for w in significant_words(denomination) if len(w) >= 4]
    if distinctive:
        present = sum(1 for w in distinctive if w in text_norm)
        if present == len(distinctive):
            score += 10
        elif present >= len(distinctive) - 1 and len(distinctive) >= 2:
            score += 4

    # Nom du dirigeant (prénom + nom) — signal moyen
    if dirigeant:
        dir_words = [w for w in normalize(dirigeant).split()
                     if len(w) >= 3 and w not in STOPWORDS]
        if len(dir_words) >= 2 and all(w in text_norm for w in dir_words):
            score += 10

    # =========================================================
    # 4) Petits bonus
    # =========================================================
    if domain.endswith(".fr"):
        score += 3
    if "mentions legales" in text_norm or "siret" in text_norm:
        score += 3

    return score


def evaluate_candidate(url: str, denomination: str, siren: str,
                       dirigeant: str, session: requests.Session) -> int:
    """Évalue un candidat en fetchant home + sous-pages si nécessaire."""
    parsed = urlparse(url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    domain = _domain_of(url)

    # 1) Homepage
    home_html = fetch_url(base + "/", session)
    if not home_html:
        return 0

    home_text = strip_html(home_html)
    score = score_from_text(home_text, denomination, siren, dirigeant, domain)

    # Match certain → on s'arrête
    if score >= DEFINITIVE_SCORE:
        return score

    # Sinon, on explore les sous-pages pour enrichir le signal
    combined_text = home_text
    for sub in VALIDATION_SUBPATHS:
        sub_url = f"{base}/{sub}"
        sub_html = fetch_url(sub_url, session)
        if sub_html:
            combined_text += " " + strip_html(sub_html)
            # re-score sur l'ensemble
            new_score = score_from_text(combined_text, denomination, siren, dirigeant, domain)
            if new_score > score:
                score = new_score
            if score >= DEFINITIVE_SCORE:
                return score
        time.sleep(random.uniform(0.3, 0.7))

    return score


def find_best_site(denomination, ville, cp, siren, dirigeant, session):
    """Retourne (url, score) du meilleur candidat validé, ou ('', 0)."""
    candidates = collect_candidates(denomination, ville, cp, session)
    if not candidates:
        return "", 0

    best_url, best_score = "", 0
    for url in candidates:
        s = evaluate_candidate(url, denomination, siren, dirigeant, session)
        if s > best_score:
            best_score = s
            parsed = urlparse(url)
            best_url = f"{parsed.scheme}://{parsed.netloc}"
        # Si on a un match certain, inutile de continuer
        if best_score >= DEFINITIVE_SCORE:
            break
        time.sleep(random.uniform(0.5, 1.0))

    if best_score >= SCORE_THRESHOLD:
        return best_url, best_score
    return "", best_score


# --- Excel I/O ------------------------------------------------------------

def read_companies(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = list(rows[0])
    data = []
    for row in rows[1:]:
        if row is None or all(v is None or v == "" for v in row):
            continue
        data.append(dict(zip(headers, row)))
    return headers, data


def write_companies(headers, companies_with_site, output_path):
    if "Site" not in headers:
        headers = headers + ["Site"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Entreprises + site"

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1B3A5C")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cf = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    bd = Border(
        left=Side("thin", "D0D0D0"), right=Side("thin", "D0D0D0"),
        top=Side("thin", "D0D0D0"), bottom=Side("thin", "D0D0D0"),
    )
    alt = PatternFill("solid", fgColor="F2F6FA")

    widths_map = {
        "N°": 6, "Dénomination": 40, "Dirigeant": 30, "SIREN": 15,
        "Forme juridique": 18, "Code postal": 14, "Ville": 25,
        "Département": 14, "Site": 40,
    }

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, bd
        ws.column_dimensions[c.column_letter].width = widths_map.get(h, 18)

    for ri, comp in enumerate(companies_with_site, 2):
        for ci, h in enumerate(headers, 1):
            v = comp.get(h, "")
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.border = cf, bd
            if ri % 2 == 0:
                c.fill = alt
            if h == "SIREN":
                c.number_format = "@"
            if h == "Site" and v:
                c.font = link_font
                c.hyperlink = v

    last_col = ws.cell(row=1, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A1:{last_col}{max(2, len(companies_with_site) + 1)}"
    ws.freeze_panes = "A2"
    wb.save(output_path)


# --- Résolution des fichiers d'entrée ------------------------------------

def resolve_input_files(args_paths):
    """Transforme les arguments CLI en liste de fichiers .xlsx réels."""
    files = []
    seen = set()
    for p in args_paths:
        matches = []
        if os.path.isdir(p):
            matches = sorted(glob.glob(os.path.join(p, "*.xlsx")))
        elif any(c in p for c in "*?[]"):
            matches = sorted(glob.glob(p))
        elif os.path.isfile(p):
            matches = [p]
        else:
            print(f"⚠️  Ignoré (introuvable) : {p}")
            continue

        for m in matches:
            # Exclut les fichiers déjà produits (pour éviter le double traitement)
            base = os.path.basename(m).lower()
            if base.endswith("_avecsite.xlsx") or "_merged_avecsite" in base:
                continue
            abspath = os.path.abspath(m)
            if abspath not in seen:
                seen.add(abspath)
                files.append(abspath)
    return files


def build_output_path(input_path: str) -> str:
    folder = os.path.dirname(os.path.abspath(input_path))
    base = os.path.basename(input_path)
    name, ext = os.path.splitext(base)
    return os.path.join(folder, f"{name}_avecsite{ext}")


def build_merged_path(first_input: str) -> str:
    folder = os.path.dirname(os.path.abspath(first_input))
    return os.path.join(folder, "entreprises_MERGED_avecsite.xlsx")


# --- Traitement d'un fichier ---------------------------------------------

def process_file(input_path, session):
    output_path = build_output_path(input_path)

    print("\n" + "=" * 60)
    print(f"  FICHIER : {os.path.basename(input_path)}")
    print("=" * 60)
    print(f"  Source : {input_path}")
    print(f"  Sortie : {output_path}\n")

    headers, companies = read_companies(input_path)
    if not companies:
        print("  ⚠️  Aucune société dans ce fichier.")
        return headers, []

    total = len(companies)
    print(f"  → {total} sociétés à traiter\n")

    results = []
    found = 0

    for idx, comp in enumerate(companies, 1):
        denom = str(comp.get("Dénomination", "") or "").strip()
        ville = str(comp.get("Ville", "") or "").strip()
        cp = str(comp.get("Code postal", "") or "").strip()
        siren = str(comp.get("SIREN", "") or "").strip()
        dirigeant = str(comp.get("Dirigeant", "") or "").strip()

        if not denom:
            continue

        # Affichage AVANT traitement pour voir l'avancement en temps réel
        print(f"  [{idx:4d}/{total}] ⏳ {denom[:52]}", flush=True)

        try:
            site, score = find_best_site(denom, ville, cp, siren, dirigeant, session)
        except Exception as e:
            print(f"  [{idx:4d}/{total}] ⚠️  {denom[:42]} — erreur : {e}", flush=True)
            site, score = "", 0

        if site:
            found += 1
            comp_copy = dict(comp)
            comp_copy["Site"] = site
            results.append(comp_copy)
            status = f"✓ [{score}]"
        else:
            status = "·     " if score == 0 else f"✗ [{score}]"

        print(f"  [{idx:4d}/{total}] {status:<10} {denom[:42]:<42} {site}", flush=True)

        if idx % SAVE_EVERY == 0 and results:
            try:
                write_companies(headers, results, output_path)
                print(f"    💾 Sauvegarde intermédiaire ({found} sites validés)", flush=True)
            except Exception as e:
                print(f"    ⚠️  Sauvegarde impossible : {e}", flush=True)

        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
        if idx % PAUSE_EVERY == 0 and idx < total:
            print(f"\n  💤 Pause {PAUSE_DURATION}s (anti rate-limit)...\n", flush=True)
            time.sleep(PAUSE_DURATION)

    # Sauvegarde finale du fichier individuel
    if results:
        write_companies(headers, results, output_path)
        print(f"\n  ✅ {found}/{total} sites validés — {os.path.basename(output_path)}")
    else:
        print(f"\n  ❌ Aucun site validé pour ce fichier.")

    return headers, results


# --- Main -----------------------------------------------------------------

def pick_files_dialog():
    """Ouvre une boîte de dialogue pour sélectionner un ou plusieurs fichiers Excel."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    # Dossier de départ : Landing RES ELEC si présent, sinon Desktop
    initial_dir = os.path.expanduser(r"~\Desktop\Landing RES ELEC")
    if not os.path.isdir(initial_dir):
        initial_dir = os.path.expanduser(r"~\Desktop")
    if not os.path.isdir(initial_dir):
        initial_dir = os.path.expanduser("~")

    paths = filedialog.askopenfilenames(
        title="Sélectionne le(s) fichier(s) Excel à enrichir (Ctrl+clic pour plusieurs)",
        initialdir=initial_dir,
        filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
    )
    root.destroy()
    return list(paths)


def main():
    # Sélection des fichiers via boîte de dialogue
    selected = pick_files_dialog()
    if not selected:
        print("❌ Aucun fichier sélectionné. Annulation.")
        sys.exit(0)

    input_files = resolve_input_files(selected)
    if not input_files:
        print("❌ Aucun fichier .xlsx valide trouvé.")
        sys.exit(1)

    print("=" * 60)
    print("  RECHERCHE DE SITES WEB — MODE MAX FIABILITÉ")
    print("=" * 60)
    print(f"  {len(input_files)} fichier(s) à traiter :")
    for f in input_files:
        print(f"    • {os.path.basename(f)}")
    print(f"  Seuil de confiance : {SCORE_THRESHOLD}")
    print(f"  Candidats max/société : {MAX_CANDIDATES_TO_VALIDATE}")
    print(f"  Sous-pages de validation : {len(VALIDATION_SUBPATHS)}")

    session = requests.Session()

    all_headers = None
    all_results = []

    for f in input_files:
        headers, results = process_file(f, session)
        if all_headers is None and headers:
            all_headers = headers
        all_results.extend(results)

    # --- Fusion finale ---
    if all_results:
        merged_path = build_merged_path(input_files[0])
        # On re-numérote les N° pour avoir une séquence propre
        for i, comp in enumerate(all_results, 1):
            if "N°" in comp:
                comp["N°"] = i
        write_companies(all_headers or list(all_results[0].keys()),
                        all_results, merged_path)
        print("\n" + "=" * 60)
        print("  FUSION TERMINÉE")
        print("=" * 60)
        print(f"  Total sites validés : {len(all_results)}")
        print(f"  Fichier fusionné    : {merged_path}")
    else:
        print("\n❌ Aucun site validé au total, aucun fichier fusionné généré.")


if __name__ == "__main__":
    main()
