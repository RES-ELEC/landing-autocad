"""
Scraper OPQIBI - TOUTES les entreprises qualifiées Électricité
Extrait : Nom, Gérant, Email, Site Internet → Excel

Usage:
    pip install requests beautifulsoup4 openpyxl
    python scrape_opqibi.py
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import time
import sys
import os

# Répertoire du script (les fichiers Excel seront créés ici)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

BASE_URL = "https://www.opqibi.com"
SEARCH_URL = (
    f"{BASE_URL}/recherche-resultat?"
    "NomSociete=&Siren=&NewRegion=&Region=&Departement="
    "&MotCle1=32&Libelle1=0&MotCle2=&Rubrique2=&Libelle2="
    "&effectif=&chiffre_affaire="
)
DELAY = 1.5
OUTPUT = os.path.join(SCRIPT_DIR, "opqibi_electricite_complet.xlsx")

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
})

# Regex pour supprimer les caractères de contrôle illégaux dans Excel/XML
ILLEGAL_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ufeff\ufffe\uffff]"
)


def clean(text):
    """Nettoie une chaîne des caractères invisibles qui font planter openpyxl."""
    if not text:
        return ""
    return ILLEGAL_CHARS_RE.sub("", text).strip()


def get_soup(url):
    resp = session.get(url, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")


def save_excel(data, filename):
    """Sauvegarde les données dans un fichier Excel formaté."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OPQIBI Electricite"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    cell_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", fgColor="D6E4F0")

    headers = ["Nom de la société", "Gérant", "Email", "Site Internet"]
    for col, titre in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=titre)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    for row, ent in enumerate(data, 2):
        for col, key in enumerate(["Nom", "Gérant", "Email", "Site Internet"], 1):
            val = clean(ent.get(key, ""))
            c = ws.cell(row=row, column=col, value=val)
            c.font, c.border = cell_font, border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if row % 2 == 0:
                c.fill = alt_fill

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.auto_filter.ref = "A1:D1"
    ws.freeze_panes = "A2"

    wb.save(filename)


def scrape_fiche(url, nom_liste):
    """Scrape une fiche et retourne un dict {Nom, Gérant, Email, Site Internet}."""
    page = get_soup(url)

    # --- Nom ---
    h1 = page.find("h1")
    nom = clean(h1.get_text(strip=True)) if h1 else clean(nom_liste)

    # --- Gérant ---
    gerant = ""
    for bold in page.find_all(["b", "strong"]):
        if "personne" in bold.get_text(strip=True).lower() and "pouvoir" in bold.get_text(strip=True).lower():
            parent = bold.parent
            if parent:
                full = parent.get_text(separator="\n", strip=True)
                lines = [l.strip() for l in full.split("\n") if l.strip()]
                found = False
                gerant_parts = []
                for line in lines:
                    if found:
                        gerant_parts.append(clean(line))
                    elif "personne" in line.lower() and "pouvoir" in line.lower():
                        found = True
                gerant = " / ".join(gerant_parts) if gerant_parts else ""
            break

    if not gerant:
        text = page.get_text()
        matches = re.findall(
            r"(Monsieur|Madame)\s+([A-ZÉÈÊËÀÂÄÙÛÜÔÖ][A-ZÉÈÊËÀÂÄÙÛÜÔÖa-zéèêëàâäùûüôö\s-]+)", text
        )
        if matches:
            gerant = " / ".join(clean(f"{m[0]} {m[1].strip()}") for m in matches)

    # --- Email ---
    email = ""
    mail_links = page.find_all("a", href=re.compile(r"^mailto:"))
    if mail_links:
        email = clean(mail_links[0].get_text(strip=True))
    if not email:
        text = page.get_text()
        mail_match = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", text)
        if mail_match:
            email = clean(mail_match.group(0))

    # --- Site Internet ---
    site = ""
    for bold in page.find_all(["b", "strong"]):
        if "site internet" in bold.get_text(strip=True).lower():
            parent = bold.parent
            if parent:
                link = parent.find_next("a", href=re.compile(r"^https?://"))
                if link and "opqibi" not in link["href"]:
                    site = clean(link.get_text(strip=True))
                    break
            nxt = bold.find_next("a", href=re.compile(r"^https?://"))
            if nxt and "opqibi" not in nxt["href"]:
                site = clean(nxt.get_text(strip=True))
            break

    if not site:
        for a in page.find_all("a", href=re.compile(r"^https?://")):
            txt = a.get_text(strip=True)
            if txt.startswith("www.") and "opqibi" not in a["href"] and "twitter" not in a["href"]:
                site = clean(txt)
                break

    return {"Nom": nom, "Gérant": gerant, "Email": email, "Site Internet": site}


# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════

print("=" * 60)
print("  SCRAPER OPQIBI - Toutes les entreprises Électricité")
print("=" * 60)

# ─── ÉTAPE 1 : Récupérer TOUS les liens ───

print("\n📋 Étape 1 : Récupération de la liste complète...")

fiche_links = []
seen_urls = set()
page_url = SEARCH_URL
page_num = 1

while page_url:
    print(f"   Page {page_num}...", end=" ")
    try:
        soup = get_soup(page_url)
    except Exception as e:
        print(f"✗ Erreur : {e}")
        break

    count_before = len(fiche_links)
    for a in soup.find_all("a", href=re.compile(r"^/fiche/")):
        name = a.get_text(strip=True)
        href = a["href"]
        full_url = BASE_URL + href
        if name and full_url not in seen_urls:
            fiche_links.append((name, full_url))
            seen_urls.add(full_url)

    new = len(fiche_links) - count_before
    print(f"{new} entreprises trouvées (total: {len(fiche_links)})")

    # Chercher la page suivante
    page_url = None
    for a in soup.find_all("a"):
        txt = a.get_text(strip=True).lower()
        href = a.get("href", "")
        if txt in ("suivant", "suivante", ">", ">>", "next", "»", "page suivante") and href:
            page_url = href if href.startswith("http") else BASE_URL + href
            break

    if not page_url:
        for a in soup.find_all("a", href=re.compile(r"page=\d+")):
            href = a.get("href", "")
            match = re.search(r"page=(\d+)", href)
            if match and int(match.group(1)) == page_num + 1:
                page_url = href if href.startswith("http") else BASE_URL + href
                break

    if new == 0 and page_num > 1:
        break

    if page_url:
        page_num += 1
        time.sleep(DELAY)

print(f"\n   → Total : {len(fiche_links)} entreprises à scraper")

if not fiche_links:
    print("\n⚠️  Aucune entreprise trouvée.")
    sys.exit(1)


# ─── ÉTAPE 2 : Scraper chaque fiche ───

total = len(fiche_links)
duree_min = int(total * DELAY / 60)
print(f"\n🔍 Étape 2 : Scraping des {total} fiches...")
print(f"   ⏱️  Durée estimée : ~{duree_min} minutes\n")

resultats = []
erreurs = []

for i, (nom_liste, url) in enumerate(fiche_links, 1):
    print(f"   [{i}/{total}] {nom_liste[:55]:<55}", end=" ")
    try:
        data = scrape_fiche(url, nom_liste)
        resultats.append(data)
        print("✓")
    except Exception as e:
        print(f"✗ {e}")
        erreurs.append((nom_liste, url, str(e)))
        resultats.append({"Nom": clean(nom_liste), "Gérant": "", "Email": "", "Site Internet": ""})

    time.sleep(DELAY)

    # Sauvegarde intermédiaire toutes les 50 entreprises
    if i % 50 == 0:
        tmp = OUTPUT.replace(".xlsx", f"_sauvegarde_{i}.xlsx")
        try:
            save_excel(resultats, tmp)
            print(f"\n   💾 Sauvegarde intermédiaire : {tmp} ({i}/{total})\n")
        except Exception as e:
            print(f"\n   ⚠️  Erreur sauvegarde intermédiaire : {e}\n")


# ─── ÉTAPE 3 : Excel final ───

print(f"\n📊 Étape 3 : Génération de {OUTPUT}...")
save_excel(resultats, OUTPUT)

print(f"\n{'=' * 60}")
print(f"  ✅ Terminé ! {len(resultats)} entreprises exportées")
print(f"  📁 Fichier : {OUTPUT}")
if erreurs:
    print(f"  ⚠️  {len(erreurs)} erreur(s) sur {total}")
print(f"{'=' * 60}")
