# -*- coding: utf-8 -*-
"""
Filtrage BET OPQIBI Electricite - v3 (scoring cumulatif)

Bareme :
  - Logiciels metier (Caneco, AutoCAD, Dialux...)   = 3 points chacun
  - "Etude electrique"                              = 2 points
  - Autres mots-cles BET elec                       = 2 points chacun

Tous les points se cumulent. Un site qui matche plusieurs fois sur
plusieurs mots-cles additionne tout.

Exemple : Caneco (3) + AutoCAD (3) + Etude electrique (2) + NFC 15-100 (2)
       = 10 points

Dependances : pip install requests beautifulsoup4 openpyxl lxml
Usage       : python filtrer_bet_electriques.py
"""
import os
import re
import sys
import time
import concurrent.futures as cf
from urllib.parse import urljoin, urlparse

import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# CONFIG
# ------------------------------------------------------------------
DOSSIER     = r"C:\Users\chapo\Desktop\landing AUTOMBE\Scrapping\scrape opqibi"
FICHIER_IN  = os.path.join(DOSSIER, "opqibi_electricite_complet.xlsx")
FICHIER_OUT = os.path.join(DOSSIER, "opqibi_electricite_CIBLES.xlsx")

SEUIL_CIBLE = 2      # score minimum pour etre considere comme cible
TIMEOUT     = 10
MAX_WORKERS = 12
MAX_PAGES   = 5
CHARS_MAX   = 250_000

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/124.0 Safari/537.36"
}

# ------------------------------------------------------------------
# BAREME
# Format : "Libelle affiche" : (regex, points)
# ------------------------------------------------------------------
BAREME = {
    # --- Logiciels metier : 3 points ---
    "Caneco":              (r"\bcaneco\b",                                   3),
    "AutoCAD":             (r"\bautocad\b",                                  3),
    "AutoCAD Electrical":  (r"\bautocad\s*electrical\b",                     3),
    "SEE Electrical":      (r"\bsee[- ]?electrical\b",                       3),
    "Dialux":              (r"\bdialux\b",                                   3),
    "Revit MEP":           (r"\brevit\s*mep\b|mep\s+[ée]lectri",             3),
    "ETAP":                (r"\betap\b",                                     3),
    "EPLAN":               (r"\beplan\b",                                    3),

    # --- Etude electrique : 2 points ---
    "Etude electrique":    (r"[ée]tudes?\s+[ée]lectriques?",                 2),

    # --- Reste BET electrique : 2 points chacun ---
    "Schema unifilaire":   (r"sch[ée]mas?\s+(?:unifilaires?|[ée]lectriques?|de\s+principe|de\s+distribution|de\s+puissance|d[' ]armoires?|de\s+c[âa]blage)", 2),
    "Note de calcul":      (r"notes?\s+de\s+calcul",                         2),
    "Bilan de puissance":  (r"bilans?\s+de\s+puissances?",                   2),
    "Selectivite":         (r"s[ée]lectivit[ée]",                            2),
    "Court-circuit / Icc": (r"courts?[\- ]circuits?|\bicc\b",                2),
    "Chute de tension":    (r"chutes?\s+de\s+tension",                       2),
    "Dimensionnement elec":(r"dimensionnements?\s+(?:[ée]lectriques?|des?\s+c[âa]bles?|des?\s+installations?\s+[ée]lectriques?)", 2),
    "NF C 15-100":         (r"nf\s*c\s*15[\- ]?100",                         2),
    "NF C 13-100":         (r"nf\s*c\s*13[\- ]?100",                         2),
    "NF C 14-100":         (r"nf\s*c\s*14[\- ]?100",                         2),
    "TGBT":                (r"\btgbt\b|tableau\s+g[ée]n[ée]ral\s+basse\s+tension", 2),
    "Armoire electrique":  (r"armoires?\s+[ée]lectriques?",                  2),
    "Plan electrique / DAO elec": (r"plans?\s+[ée]lectriques?|dao\s+[ée]lectri", 2),
    "DCE/EXE electrique":  (r"(?:dce|exe)[^.]{0,40}[ée]lectri",              2),
    "Etudes execution elec":(r"[ée]tudes?\s+d[' ]ex[ée]cution[^.]{0,40}(?:[ée]lectri|cfo|cfa|courant)", 2),
    "Photometrie":         (r"photom[ée]tri|calculs?\s+d[' ][ée]clairement", 2),
    "IRVE":                (r"\birve\b|bornes?\s+de\s+recharge",             2),
    "Genie electrique":    (r"g[ée]nie\s+[ée]lectrique",                     2),
    "CFO":                 (r"\bcfo\b",                                      2),
    "CFA":                 (r"\bcfa\b",                                      2),
    "Courant fort":        (r"courants?\s+forts?",                           2),
    "Courant faible":      (r"courants?\s+faibles?",                         2),
    "HTA/BT":              (r"hta?\s*[/\-]\s*bt",                            2),
    "Haute tension":       (r"haute\s+tension",                              2),
    "Basse tension":       (r"basse\s+tension",                              2),
    "Elec bat/tertiaire":  (r"[ée]lectricit[ée]\s+(?:du\s+)?(?:b[âa]timent|tertiaire|industriel)", 2),
    "GTB/GTC":             (r"\bgt[bc]\b|gestion\s+technique\s+(?:du\s+)?b[âa]timent", 2),
    "Desenfumage":         (r"d[ée]senfumage",                               2),
    "Paratonnerre/parafoudre": (r"paratonnerre|parafoudre",                  2),
    "SSI":                 (r"\bssi\b|s[ée]curit[ée]\s+incendie",            2),
    "Eclairage int/ext/archi": (r"[ée]clairages?\s+(?:int[ée]rieur|ext[ée]rieur|architectural|sc[ée]nique|public)", 2),
    "Photovoltaique":      (r"photovolta[iï]que",                            2),
}

PAGES_CANDIDATES = [
    "prestations", "services", "savoir-faire", "savoirfaire",
    "expertise", "expertises", "metiers", "métiers", "notre-metier",
    "competences", "compétences", "domaines", "domaines-intervention",
    "activites", "activités", "nos-activites", "nos-prestations",
    "nos-services", "nos-metiers", "ingenierie", "ingénierie",
    "electricite", "électricité", "electricity", "genie-electrique",
    "génie-électrique", "courant-fort", "cfo", "cfa",
    "references", "références", "projets", "realisations", "réalisations",
    "a-propos", "à-propos", "presentation", "présentation", "qui-sommes-nous",
]

# Compile une fois
BAREME_COMPILE = {k: (re.compile(v[0], re.IGNORECASE), v[1]) for k, v in BAREME.items()}


def normaliser_url(u):
    if not u:
        return None
    u = str(u).strip()
    if not u or u.upper() in ("NC", "NONE", "N/A"):
        return None
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    return u


def fetch(url, session):
    try:
        r = session.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        ct = r.headers.get("content-type", "")
        if r.status_code == 200 and "text/html" in ct.lower():
            return r.text
    except Exception:
        pass
    return ""


def extraire_texte(html):
    if not html:
        return ""
    try:
        soup = BeautifulSoup(html, "lxml")
    except Exception:
        soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(" ", strip=True)


def trouver_sous_pages(html, base):
    if not html:
        return []
    try:
        soup = BeautifulSoup(html, "lxml")
    except Exception:
        soup = BeautifulSoup(html, "html.parser")
    base_host = urlparse(base).netloc.replace("www.", "")
    liens = {}
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("mailto:", "tel:", "javascript:", "#")):
            continue
        full = urljoin(base, href)
        host = urlparse(full).netloc.replace("www.", "")
        if host != base_host:
            continue
        path = urlparse(full).path.lower().strip("/")
        if not path:
            continue
        texte = (a.get_text() or "").lower()
        score_url = 0
        last_seg = path.split("/")[-1]
        for cand in PAGES_CANDIDATES:
            if cand in last_seg:
                score_url += 2
            elif cand in path:
                score_url += 1
            if cand in texte:
                score_url += 1
        if score_url > 0 and full not in liens:
            liens[full] = score_url
    return [u for u, _ in sorted(liens.items(), key=lambda x: -x[1])][:MAX_PAGES - 1]


def preuve_autour(texte, rx, context=140):
    m = rx.search(texte)
    if not m:
        return ""
    s = max(0, m.start() - context)
    e = min(len(texte), m.end() + context)
    extrait = texte[s:e].replace("\n", " ").replace("\r", " ")
    extrait = re.sub(r"\s+", " ", extrait)
    return "..." + extrait + "..."


def analyser_site(nom, url_brute):
    res = {
        "nom": nom, "url": url_brute, "url_final": "",
        "score": 0, "hits": [], "detail": "",
        "preuve": "", "erreur": "", "pages_lues": 0,
    }
    url = normaliser_url(url_brute)
    if not url:
        res["erreur"] = "pas de site"
        return res

    session = requests.Session()
    html0 = fetch(url, session)
    if not html0:
        alt = (url.replace("https://www.", "https://")
               if "://www." in url
               else url.replace("https://", "https://www."))
        html0 = fetch(alt, session)
        if html0:
            url = alt
    if not html0:
        res["erreur"] = "site injoignable"
        return res

    res["url_final"] = url
    textes = [extraire_texte(html0)]
    res["pages_lues"] = 1
    for lien in trouver_sous_pages(html0, url):
        h = fetch(lien, session)
        if h:
            textes.append(extraire_texte(h))
            res["pages_lues"] += 1

    full = " ".join(textes)[:CHARS_MAX]
    full_low = full.lower()

    score_total = 0
    hits = []
    detail_parts = []
    first_preuve = ""
    for label, (rx, pts) in BAREME_COMPILE.items():
        if rx.search(full_low):
            score_total += pts
            hits.append(label)
            detail_parts.append(f"{label} (+{pts})")
            if not first_preuve:
                first_preuve = preuve_autour(full, rx)

    res["score"]  = score_total
    res["hits"]   = hits
    res["detail"] = " | ".join(detail_parts)
    res["preuve"] = first_preuve
    return res


def tier(score):
    if score >= 12:
        return "A - FORTE"
    if score >= 6:
        return "B - MOYENNE"
    if score >= SEUIL_CIBLE:
        return "C - FAIBLE"
    return "-"


def main():
    if not os.path.isfile(FICHIER_IN):
        print(f"[ERREUR] Fichier introuvable : {FICHIER_IN}")
        sys.exit(1)

    wb = openpyxl.load_workbook(FICHIER_IN)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    print(f"[*] {len(rows)} societes a analyser")
    t0 = time.time()

    results = []
    with cf.ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(analyser_site, r[0], r[3]): r for r in rows}
        for i, f in enumerate(cf.as_completed(futs), 1):
            r = futs[f]
            try:
                res = f.result()
            except Exception as e:
                res = {"nom": r[0], "url": r[3], "url_final": "",
                       "score": 0, "hits": [], "detail": "",
                       "preuve": "", "erreur": f"crash: {e}", "pages_lues": 0}
            res["gerant"] = r[1] or ""
            res["email"]  = r[2] or ""
            results.append(res)
            if res["score"] >= SEUIL_CIBLE:
                tag = f"OK[{res['score']:3}pts]"
            elif res["erreur"]:
                tag = "--       "
            else:
                tag = "         "
            nom_court = (res["nom"] or "")[:42]
            print(f"  [{i:3}/{len(rows)}] {tag} {nom_court}")

    dt = time.time() - t0
    cibles = [r for r in results if r["score"] >= SEUIL_CIBLE]
    print(f"\n[*] Termine en {dt:.0f}s")
    print(f"[*] Cibles retenues (score >= {SEUIL_CIBLE}) : {len(cibles)} / {len(results)}")
    a = sum(1 for r in cibles if r["score"] >= 12)
    b = sum(1 for r in cibles if 6 <= r["score"] < 12)
    c = sum(1 for r in cibles if SEUIL_CIBLE <= r["score"] < 6)
    print(f"    A (>=12 pts)   : {a}")
    print(f"    B (6-11 pts)   : {b}")
    print(f"    C ({SEUIL_CIBLE}-5 pts)    : {c}")

    # ----------------- Excel -----------------
    out = openpyxl.Workbook()

    # Onglet Cibles
    sh = out.active
    sh.title = "Cibles"
    entetes = ["Priorite", "Score", "Societe", "Gerant", "Email", "Site",
               "Detail des points", "Pages", "Extrait preuve"]
    sh.append(entetes)
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cibles_tries = sorted(cibles, key=lambda x: (-x["score"], x["nom"]))
    fill_a = PatternFill("solid", fgColor="C6EFCE")  # vert
    fill_b = PatternFill("solid", fgColor="FFEB9C")  # jaune
    fill_c = PatternFill("solid", fgColor="F2F2F2")  # gris
    for r in cibles_tries:
        t = tier(r["score"])
        sh.append([
            t, r["score"], r["nom"], r["gerant"], r["email"],
            r["url_final"] or r["url"],
            r["detail"], r["pages_lues"], r["preuve"],
        ])
        fill = fill_a if t.startswith("A") else fill_b if t.startswith("B") else fill_c
        for c in sh[sh.max_row]:
            c.fill = fill

    widths = [12, 8, 38, 32, 32, 32, 60, 8, 70]
    for i, w in enumerate(widths, 1):
        sh.column_dimensions[get_column_letter(i)].width = w
    for row in sh.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    sh.freeze_panes = "A2"

    # Onglet Non-cibles
    sh2 = out.create_sheet("Non-cibles")
    sh2.append(["Societe", "Site", "Score", "Erreur", "Hits partiels"])
    for r in sorted(results, key=lambda x: (-x["score"], x["nom"])):
        if r["score"] >= SEUIL_CIBLE:
            continue
        sh2.append([r["nom"], r["url_final"] or r["url"], r["score"],
                    r["erreur"], ", ".join(r["hits"])])
    for i, w in enumerate([38, 32, 8, 25, 50], 1):
        sh2.column_dimensions[get_column_letter(i)].width = w
    sh2.freeze_panes = "A2"

    # Onglet Bareme (reference)
    sh3 = out.create_sheet("Bareme")
    sh3.append(["Mot-cle / Expression", "Points"])
    for c in sh3[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
    for label, (rx, pts) in BAREME.items():
        sh3.append([label, pts])
    sh3.column_dimensions['A'].width = 35
    sh3.column_dimensions['B'].width = 10

    out.save(FICHIER_OUT)
    print(f"[OK] Fichier ecrit : {FICHIER_OUT}")


if __name__ == "__main__":
    main()
